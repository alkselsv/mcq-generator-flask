import tempfile
import csv
import openpyxl


def generate_csv(questions):

    with tempfile.NamedTemporaryFile(
        mode="w", delete=False, suffix=".csv", newline="", encoding="utf-8"
    ) as tmp:
        writer = csv.writer(tmp)

        for question in questions:
            row = [
                question["question"],
                question["options"][0],
                question["options"][1],
                question["options"][2],
                question["answer"],
                f"{question['topic_number']} {question['topic']}",
            ]
            writer.writerow(row)

        tmp_path = tmp.name

    return tmp_path


def generate_xlsx(questions):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    for row, question in enumerate(questions, 1):
        ws.cell(row=row, column=1, value=question["question"])
        ws.cell(row=row, column=2, value=question["options"][0])
        ws.cell(row=row, column=3, value=question["options"][1])
        ws.cell(row=row, column=4, value=question["options"][2])
        ws.cell(row=row, column=5, value=question["answer"])
        ws.cell(row=row, column=6, value=f"{question['topic_number']} {question['topic']}")

    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp_path = tmp.name
        wb.save(tmp_path)
    
    return tmp_path
